#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from cmdb.file_export.file_exporter import FileExporter
from cmdb.object_framework.cmdb_object_manager import object_manager
import openpyxl
import tempfile


class XlsxFileExporter(FileExporter):

    def main(self):
        workbook = self.create_xls_object()

        # save workbook
        with tempfile.NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            self.set_response(tmp.read())

    def create_xls_object(self):
        # create workbook
        workbook = openpyxl.Workbook()

        # delete default sheet
        workbook.remove(workbook.active)

        # insert data into worksheet
        run_header = True
        i = 2

        # sorts data_list by type_id

        type_id = "type_id"
        decorated = [(dict_.__dict__[type_id], dict_.__dict__) for dict_ in self.get_object_list()]
        decorated = sorted(decorated, key=lambda x: x[0], reverse=False)
        sorted_list = [dict_ for (key, dict_) in decorated]

        current_type_id = sorted_list[0][type_id]
        p = 0
        for obj in sorted_list:
            fields = obj["fields"]

            if current_type_id != obj[type_id]:
                current_type_id = obj[type_id]
                run_header = True
                i = 2

            # insert header value
            if run_header:
                # get active worksheet and rename it
                title = object_manager.get_type(obj[type_id]).label
                sheet = workbook.create_sheet(title, p)
                header = sheet.cell(row=1, column=1)
                header.value = 'public_id'
                header = sheet.cell(row=1, column=2)
                header.value = 'active'

                c = 3
                for v in fields:
                    header = sheet.cell(row=1, column=c)
                    header.value = v.get('name')
                    c = c + 1
                run_header = False

            # insert row values
            c = 3
            for key in fields:
                header = sheet.cell(row=i, column=1)
                header.value = str(obj["public_id"])
                header = sheet.cell(row=i, column=2)
                header.value = str(obj["active"])

                rows = sheet.cell(row=i, column=c)
                rows.value = str(key.get('value'))
                c = c + 1

            i = i + 1
            p = p + 1

        return workbook
