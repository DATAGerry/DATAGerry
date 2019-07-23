#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from cmdb.file_export.file_exporter import FileExporter
from flask import send_file
import openpyxl
import tempfile


class XlsFileExporter(FileExporter):

    def main(self):
        file_type = self.get_object_type()
        if file_type == 'object':
            data_list = self.get_object_by_id()
        elif file_type == 'type':
            data_list = self.get_type_by_id()
        else:
            data_list = self.get_all_objects_by_type_id()

        # create workbook
        workbook = openpyxl.Workbook()

        # insert data into worksheet
        run_header = True
        i = 2

        # sorts data_list by type_id
        sorted_list = []
        types = []

        # get type_id's from data_list
        for obj in data_list:
            if obj.type_id not in types:
                types.append(obj.type_id)
        types.sort()

        # sort data_list by types in sorted_list
        for type_id in types:
            for obj in data_list:
                if obj.type_id == type_id:
                    sorted_list.append(obj)

        current_type_id = sorted_list[0].type_id
        p = 0
        for obj in sorted_list:
            fields = obj.fields

            if current_type_id != obj.type_id:
                current_type_id = obj.type_id
                run_header = True
                i = 2

            # insert header value
            if run_header:
                # get active worksheet and rename it

                sheet = workbook.create_sheet('Mappe' + str(obj.type_id), p)
                header = sheet.cell(row=1, column=1)
                header.value = 'public_id'
                c = 2
                for v in fields:
                    header = sheet.cell(row=1, column=c)
                    header.value = v.get('name')
                    c = c + 1
                run_header = False

            # insert row values
            c = 2
            for key in fields:
                header = sheet.cell(row=i, column=1)
                header.value = str(obj.public_id)

                rows = sheet.cell(row=i, column=c)
                rows.value = str(key.get('value'))
                c = c + 1

            i = i + 1
            p = p + 1

        # save workbook
        with tempfile.NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            return send_file(tmp.name, attachment_filename="demo.xlsx")

