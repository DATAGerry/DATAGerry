# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C) 2019 NETHINKS GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from cmdb.framework.cmdb_object_manager import object_manager
from cmdb.utils import json_encoding
import csv
import io
import openpyxl
import tempfile
import json
import xml.etree.ElementTree as ET
import xml.dom.minidom

class ExportType:

    FILE_EXTENSION = None
    LABEL = None
    MULTITYPE_SUPPORT = False
    ICON = None
    DESCRIPTION = None
    ACTIVE = None

    def __init__(self):
        pass

    def export(self, object_list):
        pass


class CsvExportType(ExportType):

    FILE_EXTENSION = "csv"
    LABEL = "CSV"
    MULTITYPE_SUPPORT = False
    ICON = "fa-table"
    DESCRIPTION = "Export as CSV (only of the same type)"
    ACTIVE = True


    def __init__(self):
        pass

    def export(self, object_list):
        header = ['public_id', 'active']
        rows = []
        row = {}
        run_into = True
        public_id = object_list[0].type_id

        for obj in object_list:

            if public_id != obj.type_id:
                raise Exception({'message': 'CSV can export only object of the same type'})

            fields = obj.fields
            row.update({'public_id': str(obj.public_id), 'active': str(obj.active)})

            for key in fields:
                if run_into:
                    header.append(key.get('name'))
                row.update({key.get('name'): str(key.get('value'))})
            run_into = False
            rows.append(row)
            row = {}

        return self.csv_writer(header, rows)

    def csv_writer(self, header, rows):
        csv_file = io.StringIO()
        writer = csv.DictWriter(csv_file, fieldnames=header, dialect='excel')
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

        csv_file.seek(0)
        return csv_file


class JsonExportType(ExportType):

    FILE_EXTENSION = "json"
    LABEL = "JSON"
    MULTITYPE_SUPPORT = True
    ICON = "fa-file-text-o"
    DESCRIPTION = "Export as JSON"
    ACTIVE = True

    def export(self, object_list):
        s_keys = ['public_id', 'active', 'type_id', 'fields']
        filter_dict = []

        for obj in object_list:
            new_dict = {'public_id': 0, 'active': True, 'type': '', 'fields': []}
            for k in s_keys:
                if k in obj.__dict__:
                    if k == 'type_id':
                        new_dict.update({'type': object_manager.get_type(obj.type_id).label})
                    else:
                        new_dict.update({k: obj.__dict__[k]})

            filter_dict.append(new_dict)

        return json.dumps(filter_dict, default=json_encoding.default, indent=2)


class XlsxExportType(ExportType):

    FILE_EXTENSION = "xlsx"
    LABEL = "XLSX"
    MULTITYPE_SUPPORT = True
    ICON = "fa-file-excel-o"
    DESCRIPTION = "Export as XLS"
    ACTIVE = True

    def export(self, object_list):
        workbook = self.create_xls_object(object_list)

        # save workbook
        with tempfile.NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            return tmp.read()

    def create_xls_object(self, object_list):
        # create workbook
        workbook = openpyxl.Workbook()

        # delete default sheet
        workbook.remove(workbook.active)

        # insert data into worksheet
        run_header = True
        i = 2

        # sorts data_list by type_id

        type_id = "type_id"
        decorated = [(dict_.__dict__[type_id], dict_.__dict__) for dict_ in object_list]
        decorated = sorted(decorated, key=lambda x: x[0], reverse=False)
        sorted_list = [dict_ for (key, dict_) in decorated]

        current_type_id = sorted_list[0][type_id]
        p = 0
        for obj in sorted_list:
            fields = obj["fields"]

            if current_type_id != obj[type_id]:
                current_type_id = obj[type_id]
                run_header = True
                i = 2

            # insert header value
            if run_header:
                # get active worksheet and rename it
                title = object_manager.get_type(obj[type_id]).label
                sheet = workbook.create_sheet(title, p)
                header = sheet.cell(row=1, column=1)
                header.value = 'public_id'
                header = sheet.cell(row=1, column=2)
                header.value = 'active'

                c = 3
                for v in fields:
                    header = sheet.cell(row=1, column=c)
                    header.value = v.get('name')
                    c = c + 1
                run_header = False

            # insert row values
            c = 3
            for key in fields:
                header = sheet.cell(row=i, column=1)
                header.value = str(obj["public_id"])
                header = sheet.cell(row=i, column=2)
                header.value = str(obj["active"])

                rows = sheet.cell(row=i, column=c)
                rows.value = str(key.get('value'))
                c = c + 1

            i = i + 1
            p = p + 1

        return workbook


class XmlExportType(ExportType):

    FILE_EXTENSION = "xml"
    LABEL = "XML"
    MULTITYPE_SUPPORT = True
    ICON = "fa-code"
    DESCRIPTION = "Export as XML"
    ACTIVE = True

    def export(self, object_list):
        return self.parse_to_xml(json.loads(json.dumps(object_list, default=json_encoding.default, indent=2)))

    def parse_to_xml(self, json_obj):
        # object list
        cmdb_object_list = ET.Element('objects')

        # objects
        for obj in json_obj:
            # object
            cmdb_object = ET.SubElement(cmdb_object_list, 'object')
            cmdb_object_meta = ET.SubElement(cmdb_object, 'meta')
            # meta: public
            cmdb_object_meta_id = ET.SubElement(cmdb_object_meta, 'public_id')
            cmdb_object_meta_id.text = str(obj['public_id'])
            # meta: active
            cmdb_object_meta_active = ET.SubElement(cmdb_object_meta, 'active')
            cmdb_object_meta_active.text = str(obj['active'])
            # meta: type
            cmdb_object_meta_type = ET.SubElement(cmdb_object_meta, 'type')
            cmdb_object_meta_type.text = object_manager.get_type(obj['type_id']).label
            # fields
            cmdb_object_fields = ET.SubElement(cmdb_object, 'fields')
            for curr in obj['fields']:
                # fields: content
                field_attribs = {}
                field_attribs["name"] = str(curr['name'])
                field_attribs["value"] = str(curr['value'])
                ET.SubElement(cmdb_object_fields, "field", field_attribs)

        # return xml as string (pretty printed)
        return xml.dom.minidom.parseString(ET.tostring(cmdb_object_list, encoding='unicode', method='xml')).toprettyxml()
