# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C)  becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""TODO: document"""
import logging
import json
import re
import tempfile
import openpyxl

from cmdb.framework.exporter.format.base_exporter_format import BaseExporterFormat
from cmdb.framework.exporter.config.exporter_config_type_enum import ExporterConfigType
from cmdb.framework.rendering.render_result import RenderResult
# -------------------------------------------------------------------------------------------------------------------- #

LOGGER = logging.getLogger(__name__)

# -------------------------------------------------------------------------------------------------------------------- #
#                                               XlsxExportFormat - CLASS                                               #
# -------------------------------------------------------------------------------------------------------------------- #
class XlsxExportFormat(BaseExporterFormat):
    """TODO: ducoment"""
    FILE_EXTENSION = "xlsx"
    LABEL = "XLSX"
    MULTITYPE_SUPPORT = True
    ICON = "file-excel"
    DESCRIPTION = "Export as XLS"
    ACTIVE = True


    def export(self, data: list[RenderResult], *args):
        """Exports object_list as .xlsx file

        Args:
            data: The objects to be exported

        Returns:
            Xlsx file containing the data
        """
        workbook = self.create_xls_object(data, args)

        # save workbook
        with tempfile.NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            return tmp.read()


    def create_xls_object(self, data: list[RenderResult], args):
        """TODO: ducoment"""
        # create workbook
        workbook = openpyxl.Workbook()

        # sorts data_list by type_id
        decorated = [(dict_.type_information['type_id'], dict_) for dict_ in data]
        decorated = sorted(decorated, key=lambda x: x[0], reverse=False)
        sorted_list = [dict_ for (key, dict_) in decorated]

        # init values
        current_type_id = None
        header = ['public_id', 'active']
        columns = [] if not data else [x['name'] for x in data[0].fields]
        view = 'native'

        # Export only the shown fields chosen by the user
        if args and args[0].get("metadata", False) and \
           args[0].get('view', 'native').upper() == ExporterConfigType.RENDER.name:

            _meta = json.loads(args[0].get("metadata", ""))
            view = args[0].get('view', 'native')
            header = _meta['header']
            columns = _meta['columns']

        i = 0
        for obj in sorted_list:

            # check, if starting a new object type
            if current_type_id != obj.type_information['type_id']:
                # set current type id and fields
                current_type_id = obj.type_information['type_id']
                i = 1

                # delete default sheet
                workbook.remove(workbook.active)

                # start a new worksheet and rename it
                title = self.__normalize_sheet_title(obj.type_information['type_label'])
                sheet = workbook.create_sheet(title)

                # insert header: public_id, active
                for count, head in enumerate(header):
                    cell = sheet.cell(row=i, column=count+1)
                    cell.value = head

                # insert header: get fields from type definition
                c = len(header)
                for field in columns:
                    cell = sheet.cell(row=i, column=c+1)
                    cell.value = field
                    c = c + 1
                i = i + 1

            # insert row values: header
            for count, head in enumerate(header):
                head = 'object_id' if head == 'public_id' else head
                cell = sheet.cell(row=i, column=count+1)
                cell.value = str(obj.object_information[head])

            # get object fields as dict:
            obj_fields_dict = {}
            for field in obj.fields:
                obj_field_name = field.get('name')
                obj_fields_dict[obj_field_name] = BaseExporterFormat.summary_renderer(obj, field, view)

            # insert row values: fields
            c = len(header)+1
            for field in columns:
                cell = sheet.cell(row=i, column=c)
                cell.value = str(obj_fields_dict.get(field))
                c = c + 1

            # increase row number
            i = i + 1

        return workbook


    def __normalize_sheet_title(self, input_data):
        return re.sub('[\\*?:/\[\]]', '_', input_data)
