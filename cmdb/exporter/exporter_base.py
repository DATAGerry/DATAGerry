# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C) 2023 becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

import csv
import io
import json
import re
import tempfile
import xml.dom.minidom
import xml.etree.ElementTree as ET
import zipfile
import openpyxl

from typing import List

from cmdb.utils import json_encoding
from cmdb.utils.helpers import load_class
from cmdb.utils.system_config import SystemConfigReader
from cmdb.database.database_manager_mongo import DatabaseManagerMongo
from cmdb.exporter.exporter_utils import ExperterUtils
from cmdb.exporter.format.format_base import BaseExporterFormat
from cmdb.exporter.config.config_type import ExporterConfigType
from cmdb.framework.managers.type_manager import TypeManager
from cmdb.framework.cmdb_object_manager import CmdbObjectManager
from cmdb.framework.cmdb_render import RenderResult


database_manager = DatabaseManagerMongo(**SystemConfigReader().get_all_values_from_section('Database'))
object_manager = CmdbObjectManager(database_manager=database_manager)
type_manager = TypeManager(database_manager=database_manager)


class ZipExportType(BaseExporterFormat):

    FILE_EXTENSION = "zip"
    LABEL = "ZIP"
    MULTITYPE_SUPPORT = True
    ICON = "file-archive"
    DESCRIPTION = "Export Zipped Files"
    ACTIVE = True

    def export(self, data: List[RenderResult], *args):

        """
        Export a zip file, containing the object list sorted by type in several files.

        Args:
            data: List of objects to be exported
            args: the filetype with which the objects are stored

        Returns:
            zip file containing object files separated by types
        """

        # check what export type is requested and intitializes a new zip file in memory
        export_type = load_class(f'cmdb.exporter.exporter_base.{args[0].get("classname", "")}')()
        zipped_file = io.BytesIO()

        # Build .zip file
        with zipfile.ZipFile(zipped_file, "a", zipfile.ZIP_DEFLATED, False) as f:

            # Enters loop until the object list is empty
            while len(data) > 0:

                # Set what type the loop filters to and makes an empty list
                type_id = data[len(data) - 1].type_information['type_id']
                type_name = data[len(data) - 1].type_information['type_name']
                type_list = []

                # Filters object list to the current type_id and inserts it into type_list
                # When an object is inserted into type_list it gets deleted from object_list
                for i in range(len(data) - 1, -1, -1):
                    if data[i].type_information['type_id'] == type_id:
                        type_list.append(data[i])
                        del data[i]

                # Runs the requested export function and returns the output in the export variable
                export = export_type.export(type_list)

                # check if export output is a string, bytes or a file and inserts it into the zip file
                if isinstance(export, str) or isinstance(export, bytes):
                    f.writestr((type_name + "_ID_" + str(type_id) + "." + export_type.FILE_EXTENSION).format(i), export)
                else:
                    f.writestr((type_name + "_ID_" + str(type_id) + "." + export_type.FILE_EXTENSION).format(i), export.getvalue())

        # returns zipped file
        zipped_file.seek(0)
        return zipped_file


class CsvExportType(BaseExporterFormat):

    FILE_EXTENSION = "csv"
    LABEL = "CSV"
    MULTITYPE_SUPPORT = False
    ICON = "file-csv"
    DESCRIPTION = "Export as CSV (only of the same type)"
    ACTIVE = True

    def export(self, data: List[RenderResult], *args):

        """ Exports data as .csv file

        Args:
            data: The objects to be exported

        Returns:
            Csv file containing the data
        """

        # init values
        header = ['public_id', 'active']
        columns = [] if not data else [x['name'] for x in data[0].fields]
        rows = []
        view = 'native'
        current_type_id = None

        # Export only the shown fields chosen by the user
        if args and args[0].get("metadata", False) and args[0].get('view', 'native') == ExporterConfigType.render.name:
            _meta = json.loads(args[0].get("metadata", ""))
            view = args[0].get('view', 'native')
            header = _meta['header']
            columns = _meta['columns']

        for obj in data:
            # get type from first object and setup csv header
            if current_type_id is None:
                current_type_id = obj.type_information['type_id']

            # throw Exception if objects of different type are detected
            if current_type_id != obj.type_information['type_id']:
                raise Exception({'message': 'CSV can export only object of the same type'})

            # get object fields as dict:
            obj_fields_dict = {}
            for field in obj.fields:
                obj_field_name = field.get('name')
                obj_fields_dict[obj_field_name] = ExperterUtils.summary_renderer(obj, field, view)

            # define output row
            row = []
            for head in header:
                head = 'object_id' if head == 'public_id' else head
                row.append(str(obj.object_information[head]))
            for name in columns:
                row.append(str(obj_fields_dict.get(name, None)))
            rows.append(row)

        return self.csv_writer([*header, *columns], rows)

    def csv_writer(self, header, rows, dialect=csv.excel):

        csv_file = io.StringIO()
        writer = csv.writer(csv_file, dialect=dialect)
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)
        csv_file.seek(0)
        return csv_file


class JsonExportType(BaseExporterFormat):

    FILE_EXTENSION = "json"
    LABEL = "JSON"
    MULTITYPE_SUPPORT = True
    ICON = "file-code"
    DESCRIPTION = "Export as JSON"
    ACTIVE = True

    def export(self, data: List[RenderResult], *args):

        """Exports data as .json file

        Args:
            data: The objects to be exported

        Returns:
            Json file containing the data
        """

        # init values
        meta = False
        view = 'native'

        if args:
            meta = args[0].get("metadata", False)
            view = args[0].get('view', 'native')

        header = ['public_id', 'active', 'type_label']
        output = []

        for obj in data:
            # init columns
            columns = obj.fields

            # Export only the shown fields chosen by the user
            if meta and view == ExporterConfigType.render.name:
                _meta = json.loads(meta)
                header = _meta['header']
                columns = [x for x in columns if x['name'] in _meta['columns']]

            # init output element
            output_element = {}
            for head in header:
                head = 'object_id' if head == 'public_id' else head
                if head == 'type_label':
                    output_element.update(
                        {head: obj.type_information[head]}
                    )
                else:
                    output_element.update(
                        {head: obj.object_information[head]}
                    )
            output_element.update({'fields': []})

            # get object fields
            for field in columns:
                output_element['fields'].append({
                    'name': field.get('name'),
                    'value': ExperterUtils.summary_renderer(obj, field, view)
                })

            output.append(output_element)

        return json.dumps(output, default=json_encoding.default, ensure_ascii=False, indent=2)


class XlsxExportType(BaseExporterFormat):

    FILE_EXTENSION = "xlsx"
    LABEL = "XLSX"
    MULTITYPE_SUPPORT = True
    ICON = "file-excel"
    DESCRIPTION = "Export as XLS"
    ACTIVE = True

    def export(self, data: List[RenderResult], *args):

        """Exports object_list as .xlsx file

        Args:
            data: The objects to be exported

        Returns:
            Xlsx file containing the data
        """

        workbook = self.create_xls_object(data, args)

        # save workbook
        with tempfile.NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            return tmp.read()

    def create_xls_object(self, data: List[RenderResult], args):

        # create workbook
        workbook = openpyxl.Workbook()

        # sorts data_list by type_id
        decorated = [(dict_.type_information['type_id'], dict_) for dict_ in data]
        decorated = sorted(decorated, key=lambda x: x[0], reverse=False)
        sorted_list = [dict_ for (key, dict_) in decorated]

        # init values
        current_type_id = None
        header = ['public_id', 'active']
        columns = [] if not data else [x['name'] for x in data[0].fields]
        view = 'native'

        # Export only the shown fields chosen by the user
        if args and args[0].get("metadata", False) and args[0].get('view', 'native') == ExporterConfigType.render.name:
            _meta = json.loads(args[0].get("metadata", ""))
            view = args[0].get('view', 'native')
            header = _meta['header']
            columns = _meta['columns']

        i = 0
        for obj in sorted_list:

            # check, if starting a new object type
            if current_type_id != obj.type_information['type_id']:
                # set current type id and fields
                current_type_id = obj.type_information['type_id']
                i = 1

                # delete default sheet
                workbook.remove(workbook.active)

                # start a new worksheet and rename it
                title = self.__normalize_sheet_title(obj.type_information['type_label'])
                sheet = workbook.create_sheet(title)

                # insert header: public_id, active
                for count, head in enumerate(header):
                    cell = sheet.cell(row=i, column=count+1)
                    cell.value = head

                # insert header: get fields from type definition
                c = len(header)
                for field in columns:
                    cell = sheet.cell(row=i, column=c+1)
                    cell.value = field
                    c = c + 1
                i = i + 1

            # insert row values: header
            for count, head in enumerate(header):
                head = 'object_id' if head == 'public_id' else head
                cell = sheet.cell(row=i, column=count+1)
                cell.value = str(obj.object_information[head])

            # get object fields as dict:
            obj_fields_dict = {}
            for field in obj.fields:
                obj_field_name = field.get('name')
                obj_fields_dict[obj_field_name] = ExperterUtils.summary_renderer(obj, field, view)

            # insert row values: fields
            c = len(header)+1
            for field in columns:
                cell = sheet.cell(row=i, column=c)
                cell.value = str(obj_fields_dict.get(field))
                c = c + 1

            # increase row number
            i = i + 1

        return workbook

    def __normalize_sheet_title(self, input_data):
        return re.sub('[\\*?:/\[\]]', '_', input_data)


class XmlExportType(BaseExporterFormat):

    FILE_EXTENSION = "xml"
    LABEL = "XML"
    MULTITYPE_SUPPORT = True
    ICON = "file-alt"
    DESCRIPTION = "Export as XML"
    ACTIVE = True

    def export(self, data: List[RenderResult], *args):

        """Exports object_list as .xml file

        Args:
            data: The objects to be exported

        Returns:
            Xml file containing the data
        """

        # init values
        header = ['public_id', 'active', 'type_label']
        columns = [] if not data else [x['name'] for x in data[0].fields]
        view = 'native'

        # Export only the shown fields chosen by the user
        if args and args[0].get("metadata", False) and args[0].get('view', 'native') == ExporterConfigType.render.name:
            _meta = json.loads(args[0].get("metadata", ""))
            view = args[0].get('view', 'native')
            header = _meta['header']
            columns = _meta['columns']

        # object list
        cmdb_object_list = ET.Element('objects')

        for obj in data:
            # get object fields as dict:
            obj_fields_dict = {}
            for field in obj.fields:
                obj_field_name = field.get('name')
                obj_fields_dict[obj_field_name] = ExperterUtils.summary_renderer(obj, field, view)

            # xml output: object
            cmdb_object = ET.SubElement(cmdb_object_list, 'object')
            cmdb_object_meta = ET.SubElement(cmdb_object, 'meta')

            # xml output meta: header
            for head in header:
                head = 'object_id' if head == 'public_id' else head
                if head == 'type_label':
                    cmdb_object_meta_type = ET.SubElement(cmdb_object_meta, 'type')
                    cmdb_object_meta_type.text = obj.type_information['type_label']
                else:
                    cmdb_object_meta_id = ET.SubElement(cmdb_object_meta, head)
                    cmdb_object_meta_id.text = str(obj.object_information[head])

            # xml output: fields
            cmdb_object_fields = ET.SubElement(cmdb_object, 'fields')

            # walk over all type fields and add object field values
            for field in columns:
                field_attribs = {
                    'name': str(field),
                    'value': str(obj_fields_dict.get(field))
                }
                ET.SubElement(cmdb_object_fields, "field", field_attribs)

        # return xml as string (pretty printed)
        return xml.dom.minidom.parseString(
            ET.tostring(cmdb_object_list, encoding='unicode', method='xml')).toprettyxml()
